from io import BytesIO
from fastapi import FastAPI, File, UploadFile, Query
from fastapi.responses import JSONResponse
from openpyxl import load_workbook
from collections import deque
from fastapi.openapi.docs import get_redoc_html
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


app = FastAPI(
    title="测试API",
    description="API文档",
    version="1.0.0",
    openapi_tags=[{
        "name": "测试用户",
        "description": "API操作",
    }]
)

origins = [
    "http://127.0.0.1:3000",
    "http://0.0.0.0:3000",
    "http://localhost:3000",
    "*"
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)



def unmerge_and_process_excel(wb,in_mem_file, sheet_name=None, start_row=0, end_row=None, start_col=0, end_col=None):
    """
    读取Excel文件，拆分合并单元格内容，并返回处理后的DataFrame。
    此版本增加了对特定区域进行处理的支持。

    Args:
        excel_path (str): Excel文件的路径。
        sheet_name (str/int, optional): 要读取的sheet名称或索引。默认为活动sheet。
        start_row (int): 开始的行索引（0-based）。
        end_row (int): 结束的行索引（0-based）。
        start_col (int): 开始的列索引（0-based）。
        end_col (int): 结束的列索引（0-based）。

    Returns:
        pandas.DataFrame: 处理后的DataFrame，其中合并单元格的内容已填充到所有相应的单元格。
    """
    try:
        workbook = wb
        if sheet_name is None:
            sheet = workbook.active
        else:
            sheet = workbook[sheet_name]

        # 记录所有合并单元格及其值，但只考虑指定区域内的合并单元格
        merged_cells_info = {}
        for merged_cell_range in sheet.merged_cells:
            min_col_xl, min_row_xl, max_col_xl, max_row_xl = merged_cell_range.bounds

            # 转换为0-based索引
            min_col_pd = min_col_xl - 1
            min_row_pd = min_row_xl - 1
            max_col_pd = max_col_xl - 1
            max_row_pd = max_row_xl - 1

            # 检查合并单元格是否在指定处理区域内
            if (min_row_pd >= start_row and max_row_pd <= (end_row if end_row is not None else max_row_pd) and
                min_col_pd >= start_col and max_col_pd <= (end_col if end_col is not None else max_col_pd)):
                top_left_value = sheet.cell(row=min_row_xl, column=min_col_xl).value
                merged_cells_info[merged_cell_range.coord] = top_left_value

        # 使用pandas读取Excel的指定区域
        in_mem_file.seek(0)
        full_df = pd.read_excel(in_mem_file, sheet_name=sheet_name, header=None)

        # 根据start_row, end_row, start_col, end_col切片DataFrame
        actual_end_row = min(end_row if end_row is not None else full_df.shape[0], full_df.shape[0])
        actual_end_col = min(end_col if end_col is not None else full_df.shape[1], full_df.shape[1])

        df = full_df.iloc[start_row:actual_end_row, start_col:actual_end_col].copy()

        # 将切片后的DataFrame的索引重置，以便loc操作正确
        df.reset_index(drop=True, inplace=True)
        # 注意：这里不应该重置列索引为 range(df.shape[1])，因为这会丢失原始列的位置信息。
        # df 的列索引应该是其原始在 full_df 中的列索引，方便后续对齐合并单元格信息。
        # 如果 df 是从 full_df 切片而来，其列索引已经保留了原始信息。
        # 我们需要的是确保 df.loc 操作能够基于相对索引工作，这主要依赖于行索引。
        # 对于列索引，只要合并单元格信息转换正确，就不需要强制重置为0,1,2...

        # 遍历合并单元格信息，填充DataFrame
        for merged_range_str, value in merged_cells_info.items():
            start_cell_str, end_cell_str = merged_range_str.split(':')

            start_col_letter = ''.join(filter(str.isalpha, start_cell_str))
            start_row_num = int(''.join(filter(str.isdigit, start_cell_str)))

            end_col_letter = ''.join(filter(str.isalpha, end_cell_str))
            end_row_num = int(''.join(filter(str.isdigit, end_cell_str)))

            # 转换为0-based索引，并考虑相对切片区域的偏移量
            # 这里的转换是关键：openpyxl的索引是绝对的，我们需要将其转换为DataFrame切片后的相对索引
            rel_start_col_idx = column_index_from_string(start_col_letter) - 1 - start_col
            rel_end_col_idx = column_index_from_string(end_col_letter) - 1 - start_col
            rel_start_row_idx = start_row_num - 1 - start_row
            rel_end_row_idx = end_row_num - 1 - start_row

            # 只有当合并单元格的相对索引在当前df范围内时才进行填充
            if (rel_start_row_idx >= 0 and rel_end_row_idx < df.shape[0] and
                rel_start_col_idx >= 0 and rel_end_col_idx < df.shape[1]):
                df.loc[rel_start_row_idx:rel_end_row_idx, rel_start_col_idx:rel_end_col_idx] = value

        df = df.fillna('')  # 将所有NaN替换为空字符串

        # 移除完全为空的行和列（针对每个子表格）
        # 这里移除空行可以，但移除空列需要谨慎，因为用户可能期望保留表格内的空列
        # df.dropna(how='all', inplace=True) # 保留此行
        # df.dropna(axis=1, how='all', inplace=True) # 移除此行，因为列的裁剪在 find_table_ranges_in_sheet 中更精确地处理

        return df

    except FileNotFoundError:
        print(f"错误：文件未找到。请检查文件流。")
        return None
    except Exception as e:
        print(f"处理Excel文件时发生错误：{e}")
        return None


def dataframe_to_markdown_table(df):
    """
    将pandas DataFrame转换为Markdown表格字符串。
    """
    if df is None or df.empty:
        return ""

    headers = [str(col) for col in df.columns]

    markdown = ""

    markdown += "| " + " | ".join(headers) + " |\n"
    markdown += "| " + " | ".join(["---"] * len(headers)) + " |\n"

    for index, row in df.iterrows():
        formatted_row = []
        for item in row.values:
            cell_content = str(item).replace('|', '\\|')
            formatted_row.append(cell_content)
        markdown += "| " + " | ".join(formatted_row) + " |\n"

    return markdown


def find_table_ranges_in_sheet(wb,in_mem_file, sheet_name=None):
    """
    在Excel Sheet中查找多个潜在的表格区域，基于空行进行分隔，
    并精确确定每个表格的列范围。

    Args:
        excel_path (str): Excel文件的路径。
        sheet_name (str/int, optional): 要读取的sheet名称或索引。默认为活动sheet。

    Returns:
        list: 一个列表，每个元素是一个字典，包含'start_row', 'end_row', 'start_col', 'end_col'。
    """
    workbook = wb
    if sheet_name is None:
        sheet = workbook.active
    else:
        sheet = workbook[sheet_name]

    # 将整个sheet的数据加载到一个DataFrame中，方便处理
    # 不进行全局的行/列修剪，保留原始的布局信息
    in_mem_file.seek(0)
    full_df = pd.read_excel(in_mem_file, sheet_name=sheet_name, header=None)
    full_df = full_df.fillna('')  # 将所有NaN替换为空字符串，以便空行/列判断

    if full_df.empty:
        return []

    # 找到第一个非空单元格的行和列索引 (0-based)
    # 这将作为后续计算相对索引时的原始偏移量
    first_non_empty_row_original = (full_df.dropna(how='all').index[0] if not full_df.dropna(how='all').empty else 0)
    first_non_empty_col_original = (
        full_df.dropna(axis=1, how='all').columns[0] if not full_df.dropna(axis=1, how='all').empty else 0)

    table_ranges = []
    current_start_row = 0  # 这是 full_df 的相对行索引

    for i in range(full_df.shape[0]):
        # 检查当前行是否完全为空
        if (full_df.iloc[i] == '').all():
            # 找到一个空行，表示一个表格的结束
            if i > current_start_row:  # 确保当前数据块不为空
                # 提取当前数据块
                current_block_df = full_df.iloc[current_start_row:i, :].copy()

                # 确定此数据块的有效列范围：只对当前块进行列修剪
                valid_cols_in_block = current_block_df.loc[:, ~(current_block_df == '').all(axis=0)].columns

                if not valid_cols_in_block.empty:
                    start_col = valid_cols_in_block.min()
                    end_col = valid_cols_in_block.max()
                    table_ranges.append({
                        'start_row': current_start_row,
                        'end_row': i - 1,  # 结束行是空行的前一行
                        'start_col': start_col,
                        'end_col': end_col
                    })
            current_start_row = i + 1  # 新的表格从空行的下一行开始

    # 处理最后一个表格（如果文件末尾没有空行）
    if current_start_row < full_df.shape[0]:
        current_block_df = full_df.iloc[current_start_row:full_df.shape[0], :].copy()
        valid_cols_in_block = current_block_df.loc[:, ~(current_block_df == '').all(axis=0)].columns
        if not valid_cols_in_block.empty:
            start_col = valid_cols_in_block.min()
            end_col = valid_cols_in_block.max()
            table_ranges.append({
                'start_row': current_start_row,
                'end_row': full_df.shape[0] - 1,
                'start_col': start_col,
                'end_col': end_col
            })

    # 将识别到的相对 full_df 的索引转换为原始 Excel 的绝对索引（考虑原始前导空行空列）
    adjusted_table_ranges = []
    for r_info in table_ranges:
        adjusted_table_ranges.append({
            'start_row': r_info['start_row'] + first_non_empty_row_original,
            'end_row': r_info['end_row'] + first_non_empty_row_original,
            'start_col': r_info['start_col'] + first_non_empty_col_original,
            'end_col': r_info['end_col'] + first_non_empty_col_original
        })
    return adjusted_table_ranges






def has_border(cell, direction):
    side = getattr(cell.border, direction)
    return side and side.style is not None

def build_border_map(sheet):
    rows = sheet.max_row
    cols = sheet.max_column
    border_map = [[{'top': False, 'bottom': False, 'left': False, 'right': False}
                   for _ in range(cols)] for _ in range(rows)]

    for i in range(rows):
        for j in range(cols):
            cell = sheet.cell(row=i+1, column=j+1)
            border_map[i][j]['top'] = has_border(cell, 'top')
            border_map[i][j]['bottom'] = has_border(cell, 'bottom')
            border_map[i][j]['left'] = has_border(cell, 'left')
            border_map[i][j]['right'] = has_border(cell, 'right')
    return border_map


def find_bounding_box(area):
    rows = [r for r, c in area]
    cols = [c for r, c in area]
    return min(rows)+1, min(cols)+1, max(rows)+1, max(cols)+1

def detect_table_areas(border_map):
    visited = set()
    rows, cols = len(border_map), len(border_map[0])
    areas = []

    def is_border_cell(r, c):
        b = border_map[r][c]
        return any(b.values())

    def bfs(r, c):
        q = deque()
        q.append((r, c))
        area = {(r, c)}
        visited.add((r, c))

        while q:
            cr, cc = q.popleft()
            for dr, dc in [(-1,0),(1,0),(0,-1),(0,1)]:
                nr, nc = cr+dr, cc+dc
                if 0 <= nr < rows and 0 <= nc < cols and (nr, nc) not in visited:
                    if is_border_cell(nr, nc):
                        visited.add((nr, nc))
                        area.add((nr, nc))
                        q.append((nr, nc))
        return area

    for i in range(rows):
        for j in range(cols):
            if is_border_cell(i, j) and (i, j) not in visited:
                area = bfs(i, j)
                areas.append(find_bounding_box(area))  # 返回 (start_row, start_col, end_row, end_col)

    return areas


def extract_table_as_json(sheet, start_row, start_col, end_row, end_col, max_header_rows=3):
    table_data = []
    for r in range(start_row, end_row + 1):
        row_data = []
        for c in range(start_col, end_col + 1):
            val = sheet.cell(row=r, column=c).value
            row_data.append("" if val is None else str(val).strip())
        table_data.append(row_data)

    header_rows = 1
    for i in range(1, max_header_rows):
        if all(cell for cell in table_data[i]):
            header_rows += 1
        else:
            break

    header = []
    for col_idx in range(len(table_data[0])):
        parts = []
        for h in range(header_rows):
            parts.append(table_data[h][col_idx])
        header.append("-".join([p for p in parts if p]))

    json_data = []
    for row in table_data[header_rows:]:
        row_dict = {}
        for i in range(len(header)):
            key = header[i]
            value = row[i] if i < len(row) else ""
            row_dict[key] = value
        json_data.append(row_dict)

    return json_data

import json

def extract_all_tables_from_xlsx(wb):
    result = {}

    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        border_map = build_border_map(sheet)
        table_regions = detect_table_areas(border_map)

        result[sheetname] = []

        for idx, (r1, c1, r2, c2) in enumerate(table_regions, start=1):
            table_json = extract_table_as_json(sheet, r1, c1, r2, c2)
            result[sheetname].append({
                "region": f"R{r1}C{c1}-R{r2}C{c2}",
                "data": table_json
            })

    return result


def json_to_markdown(data: dict) -> str:
    md_lines = []
    for sheet_name, tables in data.items():
        md_lines.append(f"## {sheet_name}")
        for table in tables:
            md_lines.append(f"### 表格区域 {table['region']}")
            table_data = table['data']
            if not table_data:
                md_lines.append("*(空表格)*")
                continue
            # 表头
            headers = table_data[0].keys()
            md_lines.append("| " + " | ".join(headers) + " |")
            md_lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
            # 表数据
            for row in table_data:
                md_lines.append("| " + " | ".join(str(row[h]) for h in headers) + " |")
        md_lines.append("")  # 空行分隔
    return "\n".join(md_lines)



def get_md2(wb,in_mem_file):
    workbook = wb
    sheet_names = workbook.sheetnames
    all_markdown_outputs = []
    for sheet_name in sheet_names:
        print(f"\n===== 正在处理 Sheet 页: '{sheet_name}' =====")
        all_markdown_outputs.append(f"# Sheet: {sheet_name}\n\n")

        table_ranges = find_table_ranges_in_sheet(wb,in_mem_file, sheet_name=sheet_name)

        if not table_ranges:
            print(f"在 Sheet '{sheet_name}' 中未找到任何表格区域。")
            all_markdown_outputs.append(f"未找到任何表格区域。\n\n")
        else:
            print(f"在 Sheet '{sheet_name}' 中找到 {len(table_ranges)} 个表格区域:")
            for i, r in enumerate(table_ranges):
                print(
                    f"  表格 {i + 1}: 行 [{r['start_row']} - {r['end_row']}], 列 [{r['start_col']} - {r['end_col']}] (0-based索引)")

                # 使用找到的范围处理并打印每个表格的Markdown
                processed_df = unmerge_and_process_excel(
                    wb,
                    in_mem_file,
                    sheet_name=sheet_name,
                    start_row=r['start_row'],
                    end_row=r['end_row'] + 1,  # end_row是包含的，所以需要+1给切片
                    start_col=r['start_col'],
                    end_col=r['end_col'] + 1  # end_col是包含的，所以需要+1给切片
                )

                if processed_df is not None and not processed_df.empty:
                    # print(f"\n--- 表格 {i + 1} 处理后的DataFrame ---")
                    # print(processed_df)
                    markdown_output = dataframe_to_markdown_table(processed_df)
                    # print(f"\n--- 表格 {i + 1} 的 Markdown 输出 ---")
                    # print(markdown_output)
                    all_markdown_outputs.append(f"## 表格 {i + 1}\n")
                    all_markdown_outputs.append(markdown_output)
                    all_markdown_outputs.append("\n---\n\n")  # 每个表格之间添加分隔符
                else:
                    # print(f"表格 {i + 1} (在 {sheet_name} 中) 处理后为空，跳过。")
                    all_markdown_outputs.append(f"## 表格 {i + 1}\n")
                    all_markdown_outputs.append("处理后为空。\n\n---\n\n")

    all_markdown = "".join(all_markdown_outputs)
    res_data = "".join(all_markdown)
    return res_data



@app.post("/parse_xlsx/")
async def upload_file(
        file: UploadFile = File(...),
        type: str = Query("1", description="xlsx解析模式")

):
    """
    注意只支持xlsx(xml)文件，不支持xls(二进制)
    :param file: xlsx_file_stream
    :return: all table json
    """
    contents = await file.read()
    in_mem_file = BytesIO(contents)
    wb = load_workbook(filename=in_mem_file,data_only=True)
    all_tables = extract_all_tables_from_xlsx(wb)
    all_tables2 = get_md2(wb,in_mem_file)
    if type == "1":
        return JSONResponse(
            {
                "type": "多sheet边缘检测模式",
                "data": {
                    "json_format": json.dumps(all_tables, indent=2, ensure_ascii=False),
                    "md_format":json_to_markdown(all_tables),
                }

            }
        )
    elif type == "2":
        return JSONResponse(
            {
                "type": "多sheet单表解析模式",
                "data": {
                    "json_format": None,
                    "md_format":all_tables2,
                }
            }
        )



